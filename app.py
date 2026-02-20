# app.py
import os, io, math, datetime, time, threading
import pandas as pd
import numpy as np
from datetime import datetime as dt, timedelta
from flask import Flask, request, send_file, jsonify
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pyairtable import Table
import sys

# Force stdout to flush immediately for Render logs
sys.stdout.reconfigure(line_buffering=True)

app = Flask(__name__)

def create_google_maps_link(lat, lon):
    if (pd.isna(lat) or pd.isna(lon)) or (lat == 0 and lon == 0):
        return "NULL"
    return f'=HYPERLINK("https://www.google.com/maps?q={lat},{lon}", "View Location")'

def haversine(lat1, lon1, lat2, lon2):
    R = 6371000
    phi1 = math.radians(lat1); phi2 = math.radians(lat2)
    dphi = math.radians(lat2 - lat1); dlambda = math.radians(lon2 - lon1)
    a = math.sin(dphi/2)**2 + math.cos(phi1)*math.cos(phi2)*math.sin(dlambda/2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return R * c

def haversine_vectorized(lat1, lon1, lat2, lon2):
    """
    Vectorized haversine distance calculation using numpy.
    Returns distance in meters. Returns NaN where any input is NaN.
    """
    R = 6371000
    # Convert to radians
    phi1 = np.radians(lat1)
    phi2 = np.radians(lat2)
    dphi = np.radians(lat2 - lat1)
    dlambda = np.radians(lon2 - lon1)
    
    # Haversine formula
    a = np.sin(dphi/2)**2 + np.cos(phi1) * np.cos(phi2) * np.sin(dlambda/2)**2
    c = 2 * np.arctan2(np.sqrt(a), np.sqrt(1 - a))
    return R * c

def get_airtable_table():
    """Get Airtable table instance using environment variables."""
    api_key = os.environ.get("AIRTABLE_API_KEY")
    base_id = os.environ.get("AIRTABLE_BASE_ID")
    table_name = os.environ.get("AIRTABLE_TABLE_NAME", "Last Locations")
    
    if not api_key or not base_id:
        return None  # Airtable not configured, will fall back to old logic
    
    return Table(api_key, base_id, table_name)

def airtable_already_updated_today(table) -> bool:
    """
    Check if Airtable baseline has already been updated today (UTC) by looking at
    the Last_Baseline_Update_Date field on any record.
    """
    try:
        records = table.all(max_records=1, fields=["Last_Baseline_Update_Date"])
        if not records:
            return False
        fields = records[0].get("fields", {})
        last = fields.get("Last_Baseline_Update_Date")
        if not last:
            return False
        # Airtable may return a date string like 'YYYY-MM-DD' or full ISO with time
        if isinstance(last, str):
            date_str = last.split("T")[0]
        else:
            try:
                date_str = last.date().isoformat()
            except AttributeError:
                return False
        today_str = datetime.datetime.utcnow().date().isoformat()
        return date_str == today_str
    except Exception as e:
        print(f"[Airtable] error checking baseline date: {e}", flush=True)
        return False

def get_last_positions_from_airtable():
    """Read last reported positions from Airtable. Returns dict: {serial: {'lat': float, 'lon': float}}"""
    table = get_airtable_table()
    if not table:
        return {}
    
    try:
        # Fetch only required fields to reduce payload size
        records = table.all(fields=["Serial", "Last_Report_Lat", "Last_Report_Lon"])
        positions = {}
        for record in records:
            fields = record.get('fields', {})
            serial_raw = fields.get('Serial')
            serial = str(serial_raw).strip() if serial_raw is not None else None
            if serial:
                positions[serial] = {
                    'lat': fields.get('Last_Report_Lat'),
                    'lon': fields.get('Last_Report_Lon')
                }
        
        print(f"[Airtable] loaded {len(positions)} records", flush=True)
        
        if positions:
            k = next(iter(positions))
            print("[Airtable] sample:", k, positions[k], flush=True)
        
        bad = sum(
            1 for v in positions.values()
            if v.get("lat") in (None, "", "NULL") or v.get("lon") in (None, "", "NULL")
        )
        print(f"[Airtable] missing lat/lon: {bad}/{len(positions)}", flush=True)
        
        return positions
    except Exception as e:
        print(f"Error reading from Airtable: {e}", flush=True)
        return {}

def update_airtable_positions(current_positions):
    """
    Update Airtable with current positions. Runs in background thread.
    Uses batch operations if available, otherwise individual operations with 429 retry.
    """
    print(f"[Airtable] update requested for {len(current_positions)} trackers", flush=True)
    
    table = get_airtable_table()
    if not table:
        return
    
    print("[Airtable] table connection OK", flush=True)

    # Daily guard: only update baseline once per UTC day
    if airtable_already_updated_today(table):
        print("[Airtable] baseline already updated today; skipping update", flush=True)
        return
    
    try:
        today_str = datetime.datetime.utcnow().date().isoformat()
        # Validate and prepare positions
        validated_positions = {}
        for serial, pos in current_positions.items():
            try:
                lat = float(pos['lat'])
                lon = float(pos['lon'])
                # Validate coordinate ranges
                if not (-90 <= lat <= 90) or not (-180 <= lon <= 180):
                    print(f"Skipping invalid coordinates for {serial}: lat={lat}, lon={lon}", flush=True)
                    continue
                validated_positions[serial] = {'lat': lat, 'lon': lon}
            except (ValueError, TypeError, KeyError) as e:
                print(f"Skipping {serial} due to invalid position data: {e}", flush=True)
                continue
        
        if not validated_positions:
            return
        
        # Get all existing records to find record IDs (1 request)
        all_records = table.all(fields=["Serial"])
        record_map = {}  # {serial: record_id}
        for record in all_records:
            serial_raw = record.get('fields', {}).get('Serial')
            serial = str(serial_raw).strip() if serial_raw is not None else None
            if serial:
                record_map[serial] = record['id']
        
        # Separate updates and creates
        updates_to_do = []
        creates_to_do = []
        
        for serial, pos in validated_positions.items():
            fields = {
                'Serial': str(serial),
                'Last_Report_Lat': pos['lat'],
                'Last_Report_Lon': pos['lon'],
                'Last_Baseline_Update_Date': today_str
            }
            
            if serial in record_map:
                updates_to_do.append((record_map[serial], fields))
            else:
                creates_to_do.append(fields)
        
        # Try batch operations if available (pyairtable 2.0+)
        try:
            # Check if batch_update method exists
            if hasattr(table, 'batch_update') and updates_to_do:
                # Process updates in chunks of 10 (Airtable batch limit)
                chunk_size = 10
                for i in range(0, len(updates_to_do), chunk_size):
                    chunk = updates_to_do[i:i+chunk_size]
                    batch_data = [{'id': rid, 'fields': fields} for rid, fields in chunk]
                    try:
                        table.batch_update(batch_data)
                    except Exception as e:
                        error_str = str(e)
                        if "429" in error_str or "Rate limit" in error_str:
                            print(f"Rate limit hit during batch update, waiting 30 seconds...", flush=True)
                            time.sleep(30)
                            table.batch_update(batch_data)
                        else:
                            print(f"Error in batch update: {e}", flush=True)
                            # Fall back to individual updates for this chunk
                            for record_id, fields in chunk:
                                try:
                                    table.update(record_id, fields)
                                except Exception as update_e:
                                    print(f"Error updating record {record_id}: {update_e}", flush=True)
            else:
                # Fall back to individual updates
                for record_id, fields in updates_to_do:
                    try:
                        table.update(record_id, fields)
                    except Exception as e:
                        error_str = str(e)
                        if "429" in error_str or "Rate limit" in error_str:
                            print(f"Rate limit hit during update, waiting 30 seconds...", flush=True)
                            time.sleep(30)
                            try:
                                table.update(record_id, fields)
                            except Exception as retry_e:
                                print(f"Error retrying update for record {record_id}: {retry_e}", flush=True)
                        else:
                            print(f"Error updating record {record_id}: {e}", flush=True)
            
            # Try batch_create if available
            if hasattr(table, 'batch_create') and creates_to_do:
                chunk_size = 10
                for i in range(0, len(creates_to_do), chunk_size):
                    chunk = creates_to_do[i:i+chunk_size]
                    try:
                        table.batch_create(chunk)
                    except Exception as e:
                        error_str = str(e)
                        if "429" in error_str or "Rate limit" in error_str:
                            print(f"Rate limit hit during batch create, waiting 30 seconds...", flush=True)
                            time.sleep(30)
                            table.batch_create(chunk)
                        else:
                            print(f"Error in batch create: {e}", flush=True)
                            # Fall back to individual creates
                            for fields in chunk:
                                try:
                                    table.create(fields)
                                except Exception as create_e:
                                    print(f"Error creating record: {create_e}", flush=True)
            else:
                # Fall back to individual creates
                for fields in creates_to_do:
                    try:
                        table.create(fields)
                    except Exception as e:
                        error_str = str(e)
                        if "429" in error_str or "Rate limit" in error_str:
                            print(f"Rate limit hit during create, waiting 30 seconds...", flush=True)
                            time.sleep(30)
                            try:
                                table.create(fields)
                            except Exception as retry_e:
                                print(f"Error retrying create: {retry_e}", flush=True)
                        else:
                            print(f"Error creating record: {e}", flush=True)
        except AttributeError:
            # pyairtable version doesn't support batch operations, use individual
            for record_id, fields in updates_to_do:
                try:
                    table.update(record_id, fields)
                except Exception as e:
                    error_str = str(e)
                    if "429" in error_str or "Rate limit" in error_str:
                        print(f"Rate limit hit during update, waiting 30 seconds...", flush=True)
                        time.sleep(30)
                        try:
                            table.update(record_id, fields)
                        except Exception as retry_e:
                            print(f"Error retrying update for record {record_id}: {retry_e}", flush=True)
                    else:
                        print(f"Error updating record {record_id}: {e}", flush=True)
            
            for fields in creates_to_do:
                try:
                    table.create(fields)
                except Exception as e:
                    error_str = str(e)
                    if "429" in error_str or "Rate limit" in error_str:
                        print(f"Rate limit hit during create, waiting 30 seconds...", flush=True)
                        time.sleep(30)
                        try:
                            table.create(fields)
                        except Exception as retry_e:
                            print(f"Error retrying create: {retry_e}", flush=True)
                    else:
                        print(f"Error creating record: {e}", flush=True)
                    
    except Exception as e:
        print(f"Error updating Airtable: {e}", flush=True)

def calculate_duration(data):
    data['Timestamp'] = pd.to_datetime(data['Timestamp']).dt.tz_localize(None)
    data.sort_values(by='Timestamp', inplace=True)
    data['Location_Changed'] = (data['Address'] != data['Address'].shift(1))
    data['Location_Start'] = data['Timestamp'].where(data['Location_Changed']).ffill()
    most_recent_row = data.iloc[-1]
    duration = most_recent_row['Timestamp'] - most_recent_row['Location_Start']
    formatted_duration = f"{duration.days} days {str(duration).split(' ')[-1]}"
    return formatted_duration

def check_moved_in_24hr_vs_airtable(current_lat, current_lon, serial, airtable_positions, threshold=300):
    """
    Compare current tracker position (from data CSV) to last position stored in Airtable.
    Returns "Y" if moved > threshold meters, else "N".
    """
    if pd.isna(current_lat) or pd.isna(current_lon):
        return "N"
    
    if not serial or serial not in airtable_positions:
        return "N"  # No previous position in Airtable
    
    # Get last reported position from Airtable
    last_pos = airtable_positions[serial]
    last_lat = last_pos.get('lat')
    last_lon = last_pos.get('lon')
    
    if pd.isna(last_lat) or pd.isna(last_lon) or last_lat is None or last_lon is None:
        return "N"
    
    # Convert to float (Airtable values might be strings)
    try:
        last_lat = float(last_lat)
        last_lon = float(last_lon)
        current_lat = float(current_lat)
        current_lon = float(current_lon)
    except (ValueError, TypeError):
        return "N"  # Invalid numeric values
    
    # Calculate distance
    dist = haversine(last_lat, last_lon, current_lat, current_lon)
    return "Y" if dist > threshold else "N"

def add_filters_to_excel(file_like_bytesio):
    file_like_bytesio.seek(0)
    wb = load_workbook(file_like_bytesio)
    ws = wb.active
    max_col = ws.max_column
    max_row = ws.max_row
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out

def process_dataframes(location_data: pd.DataFrame, main_data: pd.DataFrame, date_for_name=None):
    # validate required columns
    for col in ['Serial', 'Address', 'Timestamp', 'Lat', 'Lon']:
        if col not in location_data.columns:
            raise ValueError(f"Location CSV missing required column: {col}")
    if 'Serial' not in main_data.columns:
        raise ValueError("Data CSV missing required column: Serial")
    if 'Lat' not in main_data.columns or 'Lon' not in main_data.columns:
        raise ValueError("Data CSV missing required columns: Lat and/or Lon")
    
    # Normalize serial formatting (strip whitespace, ensure string type)
    main_data["Serial"] = main_data["Serial"].astype(str).str.strip()
    location_data["Serial"] = location_data["Serial"].astype(str).str.strip()

    # Read last positions from Airtable (non-blocking: fall back to {} on error)
    try:
        airtable_positions = get_last_positions_from_airtable()
    except Exception as e:
        print(f"Error reading from Airtable, continuing without previous positions: {e}", flush=True)
        airtable_positions = {}
    
    # Confirm serial overlap
    air_serials = set(airtable_positions.keys())
    csv_serials = set(main_data["Serial"].dropna().tolist())
    overlap = air_serials.intersection(csv_serials)
    print(f"[Debug] CSV serials: {len(csv_serials)} | Airtable serials: {len(air_serials)} | Overlap: {len(overlap)}", flush=True)
    
    # Calculate durations from location_data (still needs history for Time_At_Location)
    durations = {}
    for serial, group in location_data.groupby('Serial'):
        durations[serial] = calculate_duration(group.copy())
    
    # Vectorized comparison: Build DataFrame from Airtable positions and merge
    # Convert Airtable positions dict to DataFrame for efficient merging
    if airtable_positions:
        airtable_df = pd.DataFrame([
            {
                'Serial': serial,
                'last_lat': pos.get('lat'),
                'last_lon': pos.get('lon')
            }
            for serial, pos in airtable_positions.items()
        ])
    else:
        airtable_df = pd.DataFrame(columns=['Serial', 'last_lat', 'last_lon'])
    
    # Merge Airtable positions onto main_data
    main_data = main_data.merge(airtable_df, on='Serial', how='left')
    
    # Convert Lat/Lon to numeric (coerce errors to NaN)
    main_data['Lat'] = pd.to_numeric(main_data['Lat'], errors='coerce')
    main_data['Lon'] = pd.to_numeric(main_data['Lon'], errors='coerce')
    main_data['last_lat'] = pd.to_numeric(main_data['last_lat'], errors='coerce')
    main_data['last_lon'] = pd.to_numeric(main_data['last_lon'], errors='coerce')
    
    # Vectorized distance calculation using numpy
    # Distance is NaN if any lat/lon is missing
    distance = haversine_vectorized(
        main_data['last_lat'].values,
        main_data['last_lon'].values,
        main_data['Lat'].values,
        main_data['Lon'].values
    )
    
    # Set moved_flag: "Y" if distance > 300m, else "N" (default "N" for NaN/missing)
    main_data['Moved in 24hr'] = np.where(
        (pd.notna(distance)) & (distance > 300),
        "Y",
        "N"
    )
    
    # Debug: Count how many trackers moved
    moved_flags = main_data['Moved in 24hr'].to_dict()
    y_count = sum(1 for v in moved_flags.values() if v == "Y")
    print(f"[Compare] moved=Y count: {y_count}/{len(moved_flags)}", flush=True)
    
    # Build current_positions dict for Airtable update (only valid coordinates)
    current_positions = {}
    for idx, row in main_data.iterrows():
        serial = row.get('Serial')
        if not serial:
            continue
        current_lat = row.get('Lat')
        current_lon = row.get('Lon')
        if pd.notna(current_lat) and pd.notna(current_lon):
            try:
                current_positions[serial] = {
                    'lat': float(current_lat),
                    'lon': float(current_lon)
                }
            except (ValueError, TypeError):
                continue
    
    print(f"[CSV] current_positions collected: {len(current_positions)}", flush=True)
    
    if len(current_positions) == 0:
        print("[CSV] sample raw Lat/Lon:",
              main_data[["Serial","Lat","Lon"]].head(5).to_dict("records"), flush=True)

    # Remove temporary columns used for comparison
    main_data = main_data.drop(columns=['last_lat', 'last_lon'], errors='ignore')
    
    # add outputs
    main_data['Time_At_Location'] = main_data['Serial'].map(lambda x: durations.get(x, "No data"))
    if 'Lat' in main_data.columns and 'Lon' in main_data.columns:
        main_data['Google Maps Link'] = main_data.apply(
            lambda row: create_google_maps_link(row['Lat'], row['Lon']), axis=1
        )

    # write to Excel in memory
    buf = io.BytesIO()
    main_data.to_excel(buf, index=False)
    buf.seek(0)
    buf = add_filters_to_excel(buf)

    # Update Airtable with current positions from main_data (data CSV) - asynchronously in background
    # This prevents blocking the HTTP response while Airtable updates complete
    # ONLY AFTER Excel buffer is created
    def _update_airtable():
        try:
            update_airtable_positions(current_positions)
            print("[Airtable] update finished", flush=True)
        except Exception as e:
            print(f"[Airtable] update failed: {e}", flush=True)
    
    threading.Thread(target=_update_airtable, daemon=False).start()
    print("[Airtable] update thread started", flush=True)

    # name
    if not date_for_name:
        today = datetime.datetime.now()
        date_for_name = f"{today.year}_{today.month:02d}_{today.day:02d}"
    filename = f"data_{date_for_name}_updated.xlsx"
    return filename, buf

@app.route("/process", methods=["POST"])
def process():
    """
    Accepts:
      - multipart/form-data with files:
           location_csv (file), data_csv (file)
        OR JSON with:
           location_url, data_url (public URLs Zapier can provide)
      - Optional: date_for_name (string "YYYY_MM_DD") for the output filename.
    Returns: Excel file as an attachment (application/vnd.openxmlformats-officedocument.spreadsheetml.sheet)
    """
    print("PROCESS ENDPOINT CALLED", flush=True)
    try:
        date_for_name = request.form.get("date_for_name")

        # Case 1: files uploaded directly
        if "location_csv" in request.files and "data_csv" in request.files:
            print("Incoming request received", flush=True)
            location_data = pd.read_csv(request.files["location_csv"])
            main_data = pd.read_csv(request.files["data_csv"])
            fname, buf = process_dataframes(location_data, main_data, date_for_name)
            return send_file(buf, as_attachment=True, download_name=fname,
                             mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        # Case 2: URLs provided (Zapier can pass Google Drive direct download links or other public URLs)
        payload = request.get_json(silent=True) or {}

        if not date_for_name:
            date_for_name = payload.get("date_for_name")

        if payload.get("location_url") and payload.get("data_url"):
            print("Incoming request received", flush=True)
            # Note: Render blocks outbound to Google Drive preview URLs unless they're direct-download.
            # Prefer supplying direct file content via multipart where possible.
            import requests
            loc = requests.get(payload["location_url"])
            dat = requests.get(payload["data_url"])
            loc.raise_for_status(); dat.raise_for_status()
            location_data = pd.read_csv(io.BytesIO(loc.content))
            main_data = pd.read_csv(io.BytesIO(dat.content))
            fname, buf = process_dataframes(location_data, main_data, date_for_name)
            return send_file(buf, as_attachment=True, download_name=fname,
                             mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        return jsonify({
            "error": "Provide location_csv & data_csv files, or location_url & data_url in the request body."
        }), 400

    except Exception as e:
        import traceback
        import sys
        error_msg = f"Error in /process: {str(e)}"
        print(error_msg, file=sys.stderr)
        traceback.print_exc(file=sys.stderr)
        return jsonify({
            "error": "Internal server error",
            "details": str(e)
        }), 500

@app.route("/health", methods=["GET"])
def health():
    return "ok", 200

@app.route("/debug-airtable", methods=["GET"])
def debug_airtable():
    positions = get_last_positions_from_airtable()
    sample = None
    if positions:
        k = next(iter(positions))
        sample = {"serial": k, "pos": positions[k]}
    return jsonify({"count": len(positions), "sample": sample}), 200


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5050)), debug=True)
